#!/usr/bin/env python3
"""
CLI Scraper - Store Network Control Monitoring
Login ke dashboard internal, ambil daftar toko dengan SLA < 99%,
lalu ambil detail riwayat down tiap toko.

Output bisa dalam 3 format:
  - json  (default, dicetak ke stdout - cocok untuk dibaca sistem lain)
  - csv   (disimpan ke file)
  - xlsx  (disimpan ke file Excel)

Untuk --output-format csv/xlsx, script akan menampilkan MENU INTERAKTIF
(kecuali --report-type sudah diisi) untuk memilih salah satu:
  1. dispensasi     -> satu baris per toko + detail waktu down (perilaku lama)
  2. detail-cabang  -> rekap toko SLA < 99% saja (Kode Toko, Nama Toko,
                        Provider, Status Koneksi, FD, JD, SLA), tanpa detail
                        waktu down, diurutkan SLA terkecil -> terbesar.
Kedua opsi menghasilkan FILE TERPISAH (bukan sheet dalam satu file), dan
hanya toko yang relevan yang diambil datanya (opsi detail-cabang tidak
memicu request detail down time sama sekali, jadi lebih cepat).

Usage:
    python3 sla_scraper.py --username "EDP LBK" --password "EDP123LBK" \\
        [--first-date DD-MM-YYYY] [--last-date DD-MM-YYYY] [--dc-code G157] \\
        [--output-format json|csv|xlsx] [--report-type dispensasi|detail-cabang] \\
        [--output-file NAMA_FILE]

Jika --first-date / --last-date tidak diberikan, default = hari ini.
Jika --output-format csv/xlsx dipakai tanpa --output-file, nama file
otomatis dibuat: Report_<Tipe>_<dc_code>_<firstDate>_<lastDate>.<ext>
dan disimpan langsung ke folder tetap:
    D:\\Report SLA & Data dispensasi\\
(folder dibuat otomatis kalau belum ada). Lokasi ini TIDAK mengikuti
lokasi file .exe/.py ini dijalankan -- sengaja di-fix ke path di atas
supaya tetap konsisten walau executable-nya diinstall/dipindah ke folder
lain (misal C:\\Program Files\\...\\ oleh installer aplikasi).
"""

import sys
import os
import json
import argparse
import logging
import csv
from datetime import datetime, timedelta

import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# Konfigurasi dasar
# ---------------------------------------------------------------------------
BASE_URL = "http://172.24.10.173"
LOGIN_URL = f"{BASE_URL}/Login"
CABANG_URL = f"{BASE_URL}/DashboardNasional/previewReportCabang"
DETAIL_URL = f"{BASE_URL}/DashboardNasional/previewDetailDownToko"

DEFAULT_TIMEOUT = 60  # detik (dinaikkan dari 15; query rentang tanggal panjang butuh waktu lebih lama di server)

# Logging ke stderr saja, supaya stdout tetap murni JSON
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    stream=sys.stderr,
)
logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fungsi bantu
# ---------------------------------------------------------------------------
def login(session: requests.Session, username: str, password: str) -> bool:
    """
    Melakukan POST login ke form asli:
      <form action="/Login" method="post">
        <input name="Username" .../>
        <input name="Password" .../>
      </form>
    Tidak ada anti-forgery token tersembunyi pada form ini, jadi payload
    cukup Username & Password saja.
    """
    payload = {
        "Username": username,
        "Password": password,
    }
    try:
        resp = session.post(
            LOGIN_URL,
            data=payload,
            timeout=DEFAULT_TIMEOUT,
            allow_redirects=True,
        )
        resp.raise_for_status()
        logger.info("Login request terkirim, status code: %s", resp.status_code)

        # Deteksi sederhana: jika respons akhir masih mengandung form login
        # (form action="/Login"), kemungkinan besar kredensial ditolak.
        if 'action="/Login"' in resp.text or 'id="Username"' in resp.text:
            logger.error(
                "Login sepertinya gagal (halaman login muncul kembali). "
                "Cek kembali username/password."
            )
            return False

        return True
    except requests.RequestException as e:
        logger.error("Gagal login: %s", e)
        return False


def fetch_cabang_html(session: requests.Session, dc_code: str,
                       first_date: str, last_date: str, periode: str = "DAILY",
                       type_code: str = "ALL", connection_status: str = "ALL") -> str | None:
    params = {
        "dc_code": dc_code,
        "type_code": type_code,
        "connection_status": connection_status,
        "isOnlySLA": "N",
        "periode": periode,
        "firstDate": first_date,
        "lastDate": last_date,
    }
    try:
        resp = session.get(CABANG_URL, params=params, timeout=DEFAULT_TIMEOUT)
        resp.raise_for_status()
        logger.info("Response previewReportCabang diterima, panjang HTML: %d karakter",
                     len(resp.text))
        return resp.text
    except requests.RequestException as e:
        logger.error("Gagal mengambil data cabang: %s", e)
        return None


def parse_cabang_table(html: str) -> list[dict]:
    """
    Parsing tabel utama, filter baris dengan SLA < 99.00.
    Kolom (index 0-based):
      0: store_code
      1: store_name
      2: store_ip
      3: provider      (dipakai sebagai "Type" di laporan)
      4: status_koneksi (dipakai sebagai "Sectoral" di laporan)
      5: FD  (frekuensi down)
      6: JD  (total durasi down / JDA)
      7: SLA (kolom ke-8 / indeks ke-7)
    """
    soup = BeautifulSoup(html, "html.parser")
    tbody = soup.find("tbody")
    result = []

    if tbody is None:
        logger.warning("Tidak menemukan <tbody> pada respons cabang.")
        return result

    rows = tbody.find_all("tr")
    logger.info("Total baris toko ditemukan di tabel cabang: %d", len(rows))

    for row_idx, tr in enumerate(rows, start=1):
        cols = tr.find_all("td")
        if len(cols) < 8:
            # baris tidak lengkap, lewati
            continue

        try:
            store_code = cols[0].get_text(strip=True)
            store_name = cols[1].get_text(strip=True)
            store_ip = cols[2].get_text(strip=True)
            provider = cols[3].get_text(strip=True)
            status_koneksi = cols[4].get_text(strip=True)
            fd = cols[5].get_text(strip=True)
            jd = cols[6].get_text(strip=True)
            sla_text = cols[7].get_text(strip=True)

            # FD ditulis sebagai angka asli (bukan teks), supaya Excel tidak
            # menampilkan tanda "Number Stored as Text" (segitiga hijau)
            try:
                fd = int(fd)
            except ValueError:
                pass  # biarkan apa adanya kalau memang bukan angka murni

            # Bersihkan karakter non-numerik yang umum muncul (%, spasi, dll)
            sla_clean = sla_text.replace("%", "").strip()
            sla_value = float(sla_clean)

            if sla_value < 99.00:
                result.append({
                    "store_code": store_code,
                    "store_name": store_name,
                    "store_ip": store_ip,
                    "provider": provider,
                    "status_koneksi": status_koneksi,
                    "fd": fd,
                    "jd": jd,
                    "sla": sla_value,
                })
        except (ValueError, IndexError) as e:
            logger.warning("Gagal parsing baris ke-%d: %s", row_idx, e)
            continue

    return result


def fetch_detail_down(session: requests.Session, store_code: str, store_ip: str,
                       first_date: str, last_date: str) -> list[dict]:
    """
    Ambil dan parsing tabel detail down untuk satu toko.
    Kolom:
      0: no
      1: down_awal (DD-MM-YYYY HH:mm:ss)
      2: down_akhir (DD-MM-YYYY HH:mm:ss)
      3: durasi (HH:mm:ss)
    """
    params = {
        "store_code": store_code,
        "store_ip": store_ip,
        "firstDate": first_date,
        "lastDate": last_date,
    }

    details = []
    try:
        resp = session.get(DETAIL_URL, params=params, timeout=DEFAULT_TIMEOUT)
        resp.raise_for_status()
    except requests.RequestException as e:
        logger.error("Gagal mengambil detail down untuk %s (%s): %s",
                     store_code, store_ip, e)
        return details

    try:
        soup = BeautifulSoup(resp.text, "html.parser")
        tbody = soup.find("tbody")
        if tbody is None:
            logger.warning("Tidak ada <tbody> pada detail toko %s", store_code)
            return details

        rows = tbody.find_all("tr")
        for row_idx, tr in enumerate(rows, start=1):
            cols = tr.find_all("td")
            if len(cols) < 4:
                continue
            try:
                no = cols[0].get_text(strip=True)
                down_awal = cols[1].get_text(strip=True)
                down_akhir = cols[2].get_text(strip=True)
                durasi = cols[3].get_text(strip=True)

                details.append({
                    "no": no,
                    "down_awal": down_awal,
                    "down_akhir": down_akhir,
                    "durasi": durasi,
                })
            except IndexError as e:
                logger.warning("Gagal parsing baris detail ke-%d toko %s: %s",
                               row_idx, store_code, e)
                continue
    except Exception as e:
        logger.error("Gagal parsing HTML detail toko %s: %s", store_code, e)

    return details


# ---------------------------------------------------------------------------
# Build Report Rows & Export (untuk CSV / XLSX) - format "Data Dispensasi"
# Satu baris per toko. Kolom Down Awal/Down Akhir berisi daftar multi-baris
# (satu baris teks per kejadian down, dipisah newline dalam satu sel).
# ---------------------------------------------------------------------------
REPORT_HEADERS = [
    "tanggal", "fd", "jda", "sla", "nomor",
    "down_awal", "down_akhir",
    "store_code", "store_name", "type", "sectoral",
    "kategori", "keterangan",
]


def build_report_rows(final_data: list[dict], first_date: str) -> list[dict]:
    """
    Satu baris per toko (bukan per kejadian down), sesuai format
    'Data Dispensasi': kolom Down Awal & Down Akhir berisi gabungan
    semua kejadian down (dipisah newline) dalam satu sel.
    Kolom Kategori & Keterangan sengaja dikosongkan (diisi manual nanti).
    NOMOR diisi dengan tanggal (angka hari) dari first_date.
    """
    # NOMOR = angka hari dari firstDate (format DD-MM-YYYY)
    try:
        nomor = int(first_date.split("-")[0])
    except (ValueError, IndexError):
        nomor = ""

    tanggal_display = first_date.replace("-", "/")

    rows = []
    for store in final_data:
        details = store.get("details") or []
        # Format tanggal down_awal/down_akhir diseragamkan pakai '/' (sama seperti kolom Tanggal)
        down_awal_list = [d.get("down_awal", "").replace("-", "/") for d in details]
        down_akhir_list = [d.get("down_akhir", "").replace("-", "/") for d in details]

        rows.append({
            "tanggal": tanggal_display,
            "fd": store.get("fd", ""),
            "jda": store.get("jd", ""),
            "sla": store["sla"],
            "nomor": nomor,
            "down_awal": "\n".join(down_awal_list),
            "down_akhir": "\n".join(down_akhir_list),
            "store_code": store["store_code"],
            "store_name": store["store_name"],
            "type": store.get("provider", ""),
            "sectoral": store.get("status_koneksi", ""),
            "kategori": "",
            "keterangan": "",
        })

    # Urutkan hasil dari SLA terkecil ke terbesar
    rows.sort(key=lambda r: r["sla"])

    return rows


def export_csv(rows: list[dict], filepath: str) -> None:
    header_labels = [
        "Tanggal", "FD", "JDA", "SLA", "NOMOR",
        "Down Awal", "Down Akhir",
        "Kode Toko", "Nama Toko", "Type", "Sectoral",
        "Kategori", "Keterangan",
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header_labels)
        for row in rows:
            writer.writerow([row[h] for h in REPORT_HEADERS])
    logger.info("File CSV tersimpan: %s (%d baris toko)", filepath, len(rows))


DETAIL_CABANG_HEADERS = [
    "store_code", "store_name", "provider", "status_koneksi", "fd", "jd", "sla",
]


def sort_by_sla_asc(stores: list[dict]) -> list[dict]:
    """Urutkan toko dari SLA terkecil ke terbesar (dipakai untuk Detail Cabang)."""
    return sorted(stores, key=lambda s: s["sla"])


def export_csv_detail_cabang(filtered_stores: list[dict], filepath: str) -> None:
    """
    Opsi 2 (file terpisah): rekap toko SLA < 99% mentah, satu baris per
    toko (bukan per kejadian down). Format sama seperti file 'Detail_Cabang'
    contoh: Kode Toko, Nama Toko, Provider, Status Koneksi, FD, JD, SLA.
    Diurutkan SLA terkecil -> terbesar.
    """
    rows = sort_by_sla_asc(filtered_stores)
    header_labels = [
        "Kode Toko", "Nama Toko", "Provider", "Status Koneksi", "FD", "JD", "SLA",
    ]
    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(header_labels)
        for store in rows:
            writer.writerow([store[h] for h in DETAIL_CABANG_HEADERS])
    logger.info("File CSV Detail Cabang tersimpan: %s (%d baris toko)",
                filepath, len(rows))


def _populate_detail_cabang_ws(ws, filtered_stores: list[dict]) -> None:
    """Isi satu worksheet dengan tabel Detail Cabang (SLA terkecil -> terbesar)."""
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    rows = sort_by_sla_asc(filtered_stores)

    # Judul biasa: teks hitam, tanpa background
    header_font = Font(name="Calibri", size=11, bold=True, color="FF000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Calibri", size=11)
    body_align_bottom = Alignment(vertical="bottom")
    body_align_bottom_right = Alignment(vertical="bottom", horizontal="right")  # kolom SLA

    headers = ["Kode Toko", "Nama Toko", "Provider", "Status Koneksi", "FD", "JD", "SLA"]
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = header_align

    for i, store in enumerate(rows, start=2):
        values = [store[h] for h in DETAIL_CABANG_HEADERS]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = body_align_bottom_right if col_idx == 7 else body_align_bottom

    widths = {1: 12, 2: 26, 3: 16, 4: 14, 5: 6, 6: 10, 7: 8}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"
    logger.info("Tabel Detail Cabang diisi (%d baris toko SLA < 99%%, urut SLA terkecil->terbesar).",
                len(rows))


def export_xlsx_detail_cabang(filtered_stores: list[dict], filepath: str) -> None:
    """Opsi 2 (file terpisah): file xlsx berisi HANYA sheet 'Detail Cabang'."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Detail Cabang"
    _populate_detail_cabang_ws(ws, filtered_stores)
    wb.save(filepath)
    logger.info("File XLSX Detail Cabang tersimpan: %s (%d baris toko)",
                filepath, len(filtered_stores))


def export_xlsx(rows: list[dict], filepath: str) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Data Dispensasi"


    header_font = Font(name="Calibri", size=11, bold=True, color="FF000000")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="Calibri", size=11)
    body_align_bottom = Alignment(vertical="bottom")
    body_align_bottom_wrap = Alignment(vertical="bottom", wrap_text=True)
    body_align_bottom_right = Alignment(vertical="bottom", horizontal="right")  # kolom SLA

    # ---- Header baris 1 & 2 (mirip contoh: DOWN TIME menaungi AWAL/AKHIR) ----
    single_headers_row1 = {
        1: "Tanggal", 2: "FD", 3: "JDA", 4: "SLA", 5: "NOMOR",
        8: "Kode Toko", 9: "Nama Toko", 10: "Type", 11: "Sectoral",
        12: "Kategori", 13: "Keterangan",
    }
    for col_idx, label in single_headers_row1.items():
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = header_align
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)

    # Header "DOWN TIME" menaungi kolom 6-7 (AWAL, AKHIR)
    ws.cell(row=1, column=6, value="DOWN TIME")
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
    for col_idx, label in {6: "AWAL", 7: "AKHIR"}.items():
        top = ws.cell(row=1, column=col_idx)
        top.font = header_font
        top.alignment = header_align
        sub = ws.cell(row=2, column=col_idx, value=label)
        sub.font = header_font
        sub.alignment = header_align

    # ---- Data mulai baris 3 ----
    start_row = 3
    for i, row in enumerate(rows):
        r = start_row + i
        values = [
            row["tanggal"], row["fd"], row["jda"], row["sla"], row["nomor"],
            row["down_awal"], row["down_akhir"],
            row["store_code"], row["store_name"], row["type"], row["sectoral"],
            row["kategori"], row["keterangan"],
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = body_font
            if col_idx in (6, 7):  # kolom Down Awal / Down Akhir: wrap + rata bawah
                cell.alignment = body_align_bottom_wrap
            elif col_idx == 4:  # kolom SLA: rata bawah + rata kanan
                cell.alignment = body_align_bottom_right
            else:
                cell.alignment = body_align_bottom

        # Tinggi baris menyesuaikan jumlah baris teks di kolom Down Awal
        n_lines = row["down_awal"].count("\n") + 1 if row["down_awal"] else 1
        ws.row_dimensions[r].height = max(15, n_lines * 14)

    from openpyxl.utils import get_column_letter

    # Lebar kolom
    widths = {1: 12, 2: 6, 3: 10, 4: 8, 5: 8, 6: 20, 7: 20,
              8: 10, 9: 26, 10: 14, 11: 10, 12: 22, 13: 40}
    for col_idx, w in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A3"
    wb.save(filepath)
    logger.info("File XLSX tersimpan: %s (%d baris toko)", filepath, len(rows))


def make_unique_filename(filepath: str) -> str:
    """
    Kalau filepath sudah ada, tambahkan ' (1)', ' (2)', dst sebelum ekstensi
    (mirip perilaku browser saat download file dengan nama yang sama),
    supaya file lama tidak ketimpa.
    """
    if not os.path.exists(filepath):
        return filepath

    base, ext = os.path.splitext(filepath)
    n = 1
    while True:
        candidate = f"{base} ({n}){ext}"
        if not os.path.exists(candidate):
            return candidate
        n += 1


def get_output_dir() -> str:
    """
    Folder tempat semua file laporan disimpan.

    SENGAJA di-fix ke path absolut D:\\Report SLA & Data dispensasi\\
    (bukan relatif ke lokasi script/exe ini berada), karena file
    executable-nya bisa saja diinstall/dipindah ke folder lain (misal
    C:\\Program Files\\...\\ oleh installer aplikasi Flutter), sementara
    folder hasil laporan harus tetap konsisten di satu lokasi yang sama
    supaya mudah ditemukan dan tidak butuh hak admin untuk menulis file
    (folder di luar Program Files, jadi aman dari proteksi UAC).

    Folder dibuat otomatis kalau belum ada.
    """
    output_dir = r"D:\Report SLA & Data dispensasi"
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def generate_date_range(first_date: str, last_date: str) -> list[str]:
    """Hasilkan daftar tanggal (format DD-MM-YYYY) dari first_date sampai
    last_date inklusif, satu per hari. Kalau first_date == last_date,
    hasilnya cuma 1 tanggal (perilaku sama seperti sebelumnya)."""
    start = datetime.strptime(first_date, "%d-%m-%Y")
    end = datetime.strptime(last_date, "%d-%m-%Y")
    if end < start:
        start, end = end, start
    dates = []
    current = start
    while current <= end:
        dates.append(current.strftime("%d-%m-%Y"))
        current += timedelta(days=1)
    return dates


def build_default_filename(dc_code: str, first_date: str, last_date: str, ext: str,
                            report_type: str) -> str:
    safe = lambda s: s.replace("-", "")
    # report_type yang dikirim dari main() selalu huruf kecil: "dispensasi"
    # atau "detail-cabang" (lihat prompt_report_type() & choices --report-type)
    suffix = "Dispensasi" if report_type == "dispensasi" else "Detail_Cabang"
    file_name = f"Report_{suffix}_{dc_code}_{safe(first_date)}_{safe(last_date)}.{ext}"
    full_path = os.path.join(get_output_dir(), file_name)
    return make_unique_filename(full_path)


def prompt_report_type() -> str:
    """
    Menu interaktif: minta user pilih tipe laporan sebelum data diambil.
    Return "dispensasi" atau "detail-cabang".
    """
    print("\nPilih opsi output:")
    print("  1. Rekap Dispensasi")
    print("  2. Detail Cabang")
    while True:
        choice = input("Pilihan (1/2): ").strip()
        if choice == "1":
            return "dispensasi"
        elif choice == "2":
            return "detail-cabang"
        print("Input tidak valid, masukkan 1 atau 2.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    global DEFAULT_TIMEOUT
    today_str = datetime.now().strftime("%d-%m-%Y")

    parser = argparse.ArgumentParser(
        description="Scraper SLA toko & detail down dari dashboard internal."
    )
    parser.add_argument("--first-date", default=today_str,
                         help="Format DD-MM-YYYY. Default: hari ini.")
    parser.add_argument("--last-date", default=today_str,
                         help="Format DD-MM-YYYY. Default: hari ini.")
    parser.add_argument("--dc-code", default="G157", help="Kode DC/wilayah.")
    parser.add_argument("--periode", default="DAILY",
                         choices=["DAILY", "WEEKLY", "MONTHLY", "ANNUAL"],
                         help="Periode laporan (value asli dari dropdown, bukan label tampilan). "
                              "DAILY=Harian, WEEKLY=Mingguan, MONTHLY=Bulanan, ANNUAL=Tahunan. Default: DAILY.")
    parser.add_argument("--type-code", default="ALL",
                         help="Provider filter (value dropdown Provider). Default: ALL.")
    parser.add_argument("--connection-status", default="ALL",
                         help="Status koneksi filter (ALL/UTAMA/BACKUP). Default: ALL.")
    parser.add_argument("--username", required=True, help="Username login.")
    parser.add_argument("--password", required=True, help="Password login.")
    parser.add_argument("--output-format", choices=["json", "csv", "xlsx"], default="json",
                         help="Format output. Default: json (dicetak ke stdout).")
    parser.add_argument("--report-type", choices=["dispensasi", "detail-cabang"], default=None,
                         help="Hanya berlaku untuk --output-format csv/xlsx. Pilih tipe laporan "
                              "langsung tanpa menu interaktif: 'dispensasi' (rekap + detail down "
                              "time) atau 'detail-cabang' (rekap toko SLA<99%% tanpa detail down, "
                              "diurutkan SLA terkecil->terbesar). Jika kosong, akan muncul menu "
                              "pilihan interaktif saat script dijalankan.")
    parser.add_argument("--output-file", default=None,
                         help="Path file output untuk csv/xlsx. Jika kosong, nama dibuat otomatis.")
    parser.add_argument("--timeout", type=int, default=DEFAULT_TIMEOUT,
                         help=f"Timeout HTTP request dalam detik. Default: {DEFAULT_TIMEOUT}. "
                              "Naikkan kalau rentang tanggal panjang (misal >7 hari) dan sering timeout.")
    args = parser.parse_args()

    # Terapkan timeout dari CLI (override default module-level)
    DEFAULT_TIMEOUT = args.timeout

    # Peringatan: dashboard sumber tampaknya hanya menyimpan riwayat detail
    # down (previewDetailDownToko) untuk 4 hari terakhir.
    try:
        first_date_obj = datetime.strptime(args.first_date, "%d-%m-%Y")
        days_ago = (datetime.now() - first_date_obj).days
        if days_ago > 4:
            logger.warning(
                "Tanggal %s lebih dari 4 hari lalu - riwayat Down Awal/Akhir mungkin sudah tidak tersedia.",
                args.first_date,
            )
    except ValueError:
        pass  # format tanggal akan divalidasi/gagal secara alami di request berikutnya

    session = requests.Session()

    # Langkah 1: Login
    if not login(session, args.username, args.password):
        logger.error("Proses dihentikan karena login gagal.")
        print(json.dumps([]))  # tetap keluarkan JSON valid meski kosong
        sys.exit(1)

    # Untuk csv/xlsx, tentukan tipe laporan (menu interaktif jika belum dipilih
    # lewat --report-type). Untuk json, perilaku lama tetap dipakai (data lengkap).
    report_type = None
    if args.output_format in ("csv", "xlsx"):
        report_type = args.report_type or prompt_report_type()
        logger.info("Tipe laporan dipilih: %s", report_type)

    # ---- Opsi 1: Rekap Dispensasi -> proses PER TANGGAL (bukan gabungan) ----
    # Supaya data down setiap toko tidak tercampur lintas tanggal, dan
    # FD/JDA/SLA akurat sesuai tanggal masing-masing, kita ulang proses
    # fetch cabang + detail down untuk tiap hari dalam rentang firstDate..lastDate.
    if report_type == "dispensasi":
        date_list = generate_date_range(args.first_date, args.last_date)
        logger.info("Memproses %d tanggal secara berurutan: %s",
                     len(date_list), ", ".join(date_list))

        all_rows = []
        for day in date_list:
            logger.info("=== Memproses tanggal %s ===", day)
            day_html = fetch_cabang_html(session, args.dc_code, day, day,
                                          args.periode, args.type_code, args.connection_status)
            if day_html is None:
                logger.warning("Lewati tanggal %s karena gagal mengambil data cabang.", day)
                continue

            day_filtered = parse_cabang_table(day_html)
            logger.info("Tanggal %s: ditemukan %d toko dengan SLA < 99%%.", day, len(day_filtered))

            day_final_data = []
            skipped_empty = 0
            for store in day_filtered:
                store_code = store["store_code"]
                store_ip = store["store_ip"]
                try:
                    details = fetch_detail_down(session, store_code, store_ip, day, day)
                except Exception as e:
                    logger.error("Error tak terduga saat proses toko %s tanggal %s: %s",
                                  store_code, day, e)
                    details = []

                if not details:
                    # Riwayat down tidak tersedia (kemungkinan sudah lewat batas
                    # retensi server) - lewati toko ini untuk tanggal ini,
                    # supaya laporan hanya berisi baris yang ada bukti down time-nya.
                    skipped_empty += 1
                    continue

                day_final_data.append({
                    "store_code": store_code,
                    "store_name": store["store_name"],
                    "store_ip": store_ip,
                    "provider": store.get("provider", ""),
                    "status_koneksi": store.get("status_koneksi", ""),
                    "fd": store.get("fd", ""),
                    "jd": store.get("jd", ""),
                    "sla": store["sla"],
                    "details": details,
                })

            if skipped_empty:
                logger.info("Tanggal %s: %d toko dilewati (riwayat down time tidak tersedia).",
                             day, skipped_empty)

            day_rows = build_report_rows(day_final_data, day)
            all_rows.extend(day_rows)

        ext = args.output_format
        filepath = args.output_file or build_default_filename(
            args.dc_code, args.first_date, args.last_date, ext, report_type
        )
        try:
            if ext == "csv":
                export_csv(all_rows, filepath)
            else:
                export_xlsx(all_rows, filepath)
        except Exception as e:
            logger.error("Gagal menulis file output (%s): %s", filepath, e)
            print(json.dumps(all_rows, ensure_ascii=False))
            sys.exit(1)
        print(f"OK - file Rekap Dispensasi tersimpan di: {filepath}")
        return

    # ---- Opsi 2 (Detail Cabang) & mode JSON lama: tetap 1 kali fetch gabungan ----
    # Langkah 2: Ambil daftar toko
    html = fetch_cabang_html(session, args.dc_code, args.first_date, args.last_date,
                              args.periode, args.type_code, args.connection_status)
    if html is None:
        logger.error("Proses dihentikan karena gagal mengambil data cabang.")
        print(json.dumps([]))
        sys.exit(1)

    # Langkah 3: Filter SLA < 99%
    filtered_stores = parse_cabang_table(html)
    logger.info("Ditemukan %d toko dengan SLA < 99%%.", len(filtered_stores))

    if len(filtered_stores) == 0:
        debug_path = "debug_response_cabang.html"
        try:
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(html)
            logger.info(
                "Hasil 0 toko - HTML mentah dari server disimpan ke '%s' untuk pengecekan manual.",
                debug_path,
            )
        except Exception as e:
            logger.warning("Gagal menyimpan file debug: %s", e)

    # ---- Opsi 2: Detail Cabang -> tidak perlu ambil detail down sama sekali ----
    if report_type == "detail-cabang":
        ext = args.output_format
        filepath = args.output_file or build_default_filename(
            args.dc_code, args.first_date, args.last_date, ext, report_type
        )
        try:
            if ext == "csv":
                export_csv_detail_cabang(filtered_stores, filepath)
            else:
                export_xlsx_detail_cabang(filtered_stores, filepath)
        except Exception as e:
            logger.error("Gagal menulis file output (%s): %s", filepath, e)
            print(json.dumps(filtered_stores, ensure_ascii=False))
            sys.exit(1)
        print(f"OK - file Detail Cabang tersimpan di: {filepath}")
        return

    # Langkah 4 & 5: Ambil detail down per toko (sekuensial, tahan error)
    # -> hanya dijalankan untuk opsi Dispensasi (dan mode json lama)
    final_data = []
    for store in filtered_stores:
        store_code = store["store_code"]
        store_ip = store["store_ip"]
        try:
            details = fetch_detail_down(
                session, store_code, store_ip,
                args.first_date, args.last_date
            )
        except Exception as e:
            logger.error("Error tak terduga saat proses toko %s: %s", store_code, e)
            details = []

        final_data.append({
            "store_code": store_code,
            "store_name": store["store_name"],
            "store_ip": store_ip,
            "provider": store.get("provider", ""),
            "status_koneksi": store.get("status_koneksi", ""),
            "fd": store.get("fd", ""),
            "jd": store.get("jd", ""),
            "sla": store["sla"],
            "details": details,
        })

    # Cek konkret: kalau data down yang benar-benar didapat ternyata mulai
    # lebih baru dari firstDate yang diminta, berarti datanya kepotong oleh
    # batas retensi server (bukan cuma dugaan berdasarkan tanggal hari ini).
    try:
        requested_start = datetime.strptime(args.first_date, "%d-%m-%Y")
        earliest_found = None
        for store in final_data:
            for d in store["details"]:
                try:
                    dt = datetime.strptime(d["down_awal"], "%d-%m-%Y %H:%M:%S")
                except (ValueError, KeyError):
                    continue
                if earliest_found is None or dt < earliest_found:
                    earliest_found = dt

        if earliest_found is not None and earliest_found.date() > requested_start.date():
            logger.warning(
                "Data down tersedia mulai %s, bukan dari %s yang diminta - riwayat sebelumnya sudah tidak ada.",
                earliest_found.strftime("%d-%m-%Y"), args.first_date,
            )
    except ValueError:
        pass

    # Langkah 6: Output (hanya tercapai untuk mode JSON; csv/xlsx sudah
    # 'return' lebih awal di masing-masing branch report_type di atas)
    print(json.dumps(final_data, ensure_ascii=False))


if __name__ == "__main__":
    main()